# S.T.A.R. Backend - FastAPI Main Application
import multiprocessing
import sys
import os
import webbrowser
import threading
from fastapi import FastAPI, Depends, HTTPException, status, Request
from fastapi.security import OAuth2PasswordBearer
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime, timedelta
import io
import csv
import traceback
from sqlalchemy import text, func

# Fix for console-less mode (PyInstaller with console=False)
# When running without console, sys.stdout/stderr are None which crashes uvicorn logging
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w')
if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')

# Required for PyInstaller Windows executable
if __name__ == "__main__":
    multiprocessing.freeze_support()


from app.database import get_db, init_database

from app.models import SevaCatalog, SystemSetting, AuditLog, DailyPanchang
from app.schemas import (
    UserCreate, UserResponse, Token, UserLogin, TokenData,
    SevaResponse, SevaUpdate, SevaCreate,
    TransactionCreate, TransactionResponse,
    ShaswataCreate, ShaswataResponse, ShaswataDispatchAdhoc,
    PasswordChange, SystemSettingResponse, SystemSettingUpdate, AuditLogResponse,
)

from app.crud import (
    create_user, get_user_by_username,
    get_all_sevas, create_seva_fn, update_seva,
    book_seva_transaction,
    get_today_transactions, get_daily_transactions, get_daily_stats,
    create_shaswata_subscription, get_shaswata_subscriptions,
    get_daily_summary, get_transaction_trends,
    get_financial_report, get_enhanced_report, get_collection_details,
    log_dispatch, log_feedback_sent, get_pending_feedback_subscriptions,
    send_address_confirmation, confirm_devotee_address, reset_address_confirmation
)

from app.panchang import PanchangCalculator
from app import daiva_setu  # Genesis Protocol (Level 15)
from app.sync_engine import sync_engine
from app.shaswata_service import (
    populate_upcoming_events, get_upcoming_events,
    mark_event_dispatched, record_delivery_feedback, dispatch_adhoc_event,
    get_pending_delivery_checks, get_communication_history,
    MessagingService
)

# Authentication Imports
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import timedelta
from app.models import User

# =============================================================================
# FastAPI Application Setup
# =============================================================================

app = FastAPI(
    title="S.T.A.R. API (Dev Mode)",
    description="Subramanya Temple App & Registry - Backend API [v0.0.00]",
    version="0.0.00"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include Routers (if split)
# app.include_router(users.router)

@app.on_event("startup")
def startup_event():
    """Initialize database tables on app startup."""
    init_database()
    sync_engine.start()

@app.on_event("shutdown")
def shutdown_event():
    sync_engine.stop()


# Register Genesis Protocol Router (AI Engine)
app.include_router(daiva_setu.router)

# =============================================================================
# Authentication Configuration & Helpers
# =============================================================================

# SECRET_KEY should be kept secret in production!
SECRET_KEY = "supersecretkey_change_this_for_production"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 hours

pwd_context = CryptContext(schemes=["bcrypt", "pbkdf2_sha256"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError:
        raise credentials_exception
        
    user = db.query(User).filter(User.username == token_data.username).first()
    if user is None:
        raise credentials_exception
    return user

# =============================================================================
# Auth Routes
# =============================================================================

@app.post("/create-admin", response_model=Token, tags=["Authentication"])
def create_initial_admin(db: Session = Depends(get_db)):
    """One-time setup endpoint to create the admin user"""
    # Check if admin already exists
    user = db.query(User).filter(User.username == "admin").first()
    if user:
        raise HTTPException(status_code=400, detail="Admin user already exists")
    
    hashed_pwd = get_password_hash("password123")
    new_user = User(username="admin", hashed_password=hashed_pwd, role="admin")
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    # Generate token
    access_token = create_access_token(data={"sub": new_user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/token", response_model=Token, tags=["Authentication"])
def login_for_access_token(form_data: UserLogin, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

# =============================================================================
# API Routes - User Management
# =============================================================================

@app.get("/users/me", response_model=UserResponse, tags=["User Management"])
def read_users_me(current_user: User = Depends(get_current_user)):
    """Get current user details"""
    return current_user

@app.get("/users", response_model=List[UserResponse], tags=["User Management"])
def list_users(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List all users (Admin only)"""
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can view users"
        )
    return db.query(User).all()

@app.post("/users", response_model=UserResponse, tags=["User Management"])
def create_user(
    user_data: UserCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new user (Admin only)"""
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can create users"
        )
    
    # Check if username already exists
    existing = db.query(User).filter(User.username == user_data.username).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already exists"
        )
    
    # Create user
    hashed_pwd = get_password_hash(user_data.password)
    new_user = User(
        username=user_data.username,
        hashed_password=hashed_pwd,
        role=user_data.role.lower()
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.delete("/users/{user_id}", tags=["User Management"])
def delete_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a user (Admin only)"""
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can delete users"
        )
    
    # Find user
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete yourself"
        )
    
    db.delete(user)
    db.commit()
    return {"message": f"User '{user.username}' deleted successfully"}


@app.put("/users/me/password", tags=["User Management"])
def change_my_password(
    data: PasswordChange,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Change the current user's password"""
    # Verify current password
    if not verify_password(data.current_password, current_user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Update password
    current_user.hashed_password = get_password_hash(data.new_password)
    db.commit()
    
    # Log the action
    audit = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action="UPDATE",
        resource_type="USER",
        resource_id=str(current_user.id),
        details='{"field": "password", "note": "Password changed by user"}'
    )
    db.add(audit)
    db.commit()
    
    return {"message": "Password updated successfully"}


# =============================================================================
# API Routes - System Settings
# =============================================================================

@app.get("/settings", response_model=List[SystemSettingResponse], tags=["Settings"])
def get_all_settings(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all system settings"""
    return db.query(SystemSetting).order_by(SystemSetting.category, SystemSetting.key).all()


@app.put("/settings", tags=["Settings"])
def update_settings(
    data: SystemSettingUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Batch update system settings (Admin only)"""
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can update settings"
        )
    
    updated_keys = []
    for key, value in data.settings.items():
        setting = db.query(SystemSetting).filter(SystemSetting.key == key).first()
        if setting:
            old_value = setting.value
            setting.value = str(value)
            updated_keys.append(key)
        else:
            # Create new setting if it doesn't exist
            new_setting = SystemSetting(key=key, value=str(value))
            db.add(new_setting)
            updated_keys.append(key)
    
    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action="UPDATE",
        resource_type="SETTING",
        details=f'{{"updated_keys": {updated_keys}}}'
    )
    db.add(audit)
    db.commit()
    
    return {"message": f"Updated {len(updated_keys)} settings", "updated_keys": updated_keys}


# =============================================================================
# API Routes - System Operations
# =============================================================================

@app.get("/system/backup", tags=["System"])
def download_backup(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download the SQLite database file as backup (Admin only)"""
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can download backups"
        )
    
    # Determine DB path from database.py
    from app.database import DATABASE_PATH as DB_PATH
    
    if not os.path.exists(DB_PATH):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Database file not found"
        )
    
    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        username=current_user.username,
        action="BACKUP",
        resource_type="DATABASE",
        details='{"action": "Database backup downloaded"}'
    )
    db.add(audit)
    db.commit()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"star_temple_backup_{timestamp}.db"
    
    return FileResponse(
        path=DB_PATH,
        filename=filename,
        media_type="application/octet-stream"
    )


@app.get("/system/health", tags=["System"])
def system_health(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get system health stats: DB size, record counts, uptime"""
    from app.database import DATABASE_PATH as DB_PATH
    from app.models import Devotee, ShaswataSubscription, Transaction
    
    # DB file size
    db_size_bytes = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    db_size_mb = round(db_size_bytes / (1024 * 1024), 2)
    
    # Record counts
    total_sevas = db.query(func.count(SevaCatalog.id)).scalar() or 0
    active_sevas = db.query(func.count(SevaCatalog.id)).filter(SevaCatalog.is_active == True).scalar() or 0
    total_devotees = db.query(func.count(Devotee.id)).scalar() or 0
    total_subscriptions = db.query(func.count(ShaswataSubscription.id)).scalar() or 0
    total_transactions = db.query(func.count(Transaction.id)).scalar() or 0
    total_users = db.query(func.count(User.id)).scalar() or 0
    
    return {
        "database": {
            "size_mb": db_size_mb,
            "size_bytes": db_size_bytes,
            "path": str(DB_PATH)
        },
        "records": {
            "total_sevas": total_sevas,
            "active_sevas": active_sevas,
            "total_devotees": total_devotees,
            "total_subscriptions": total_subscriptions,
            "total_transactions": total_transactions,
            "total_users": total_users
        },
        "server": {
            "version": "0.0.00",
            "status": "running"
        }
    }


# =============================================================================
# API Routes - Audit Logs
# =============================================================================

@app.get("/audit-logs", response_model=List[AuditLogResponse], tags=["Audit"])
def get_audit_logs(
    limit: int = 50,
    offset: int = 0,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get paginated audit logs (Admin only)"""
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can view audit logs"
        )
    
    query = db.query(AuditLog)
    
    if action:
        query = query.filter(AuditLog.action == action.upper())
    if resource_type:
        query = query.filter(AuditLog.resource_type == resource_type.upper())
    
    return query.order_by(AuditLog.timestamp.desc()).offset(offset).limit(limit).all()

# =============================================================================
# API Routes - Health & Sevas
# =============================================================================

@app.get("/", tags=["Health"])
def root():
    # In production (bundled exe), serve the React app
    if getattr(sys, 'frozen', False):
        base_dir = sys._MEIPASS
        index_file = os.path.join(base_dir, 'static', 'index.html')
        if os.path.exists(index_file):
            return FileResponse(index_file)
    # In dev mode or if no static files, return API status
    return {"message": "S.T.A.R. API is online"}

@app.get("/sevas", response_model=List[SevaResponse], tags=["Seva Catalog"])
def get_all_sevas(active_only: bool = True, db: Session = Depends(get_db)):
    query = db.query(SevaCatalog)
    if active_only:
        query = query.filter(SevaCatalog.is_active == True)
    return query.all()

@app.get("/sevas/{seva_id}", response_model=SevaResponse, tags=["Seva Catalog"])
def get_seva_by_id(seva_id: int, db: Session = Depends(get_db)):
    seva = db.query(SevaCatalog).filter(SevaCatalog.id == seva_id).first()
    if seva is None:
        raise HTTPException(status_code=404, detail="Seva not found")
    return seva

@app.put("/sevas/{seva_id}", response_model=SevaResponse, tags=["Seva Catalog"])
def update_seva_endpoint(
    seva_id: int, 
    seva_update: SevaUpdate, 
    current_user: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """
    Update Seva details (Admin only).
    """
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can update Seva details"
        )
    
    try:
        updated_seva = update_seva(db, seva_id, seva_update)
        if not updated_seva:
             raise HTTPException(status_code=404, detail="Seva not found")
        return updated_seva
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update Seva: {str(e)}")

@app.post("/sevas", response_model=SevaResponse, status_code=status.HTTP_201_CREATED, tags=["Seva Catalog"])
def create_seva_endpoint(
    seva: SevaCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Create a new Seva (Admin only).
    """
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can create new Sevas"
        )
    
    try:
        return create_seva_fn(db, seva)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create Seva: {str(e)}")


@app.delete("/sevas/{seva_id}", status_code=status.HTTP_204_NO_CONTENT, tags=["Seva Catalog"])
def delete_seva_endpoint(
    seva_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Soft delete a Seva (Admin only).
    """
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can delete Sevas"
        )
    
    from app.crud import delete_seva
    updated_seva = delete_seva(db, seva_id)
    if not updated_seva:
        raise HTTPException(status_code=404, detail="Seva not found")
    return None


@app.post("/sevas/{seva_id}/restore", response_model=SevaResponse, tags=["Seva Catalog"])
def restore_seva_endpoint(
    seva_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Restore a soft-deleted Seva (Admin only).
    """
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can restore Sevas"
        )
    
    from app.crud import restore_seva
    updated_seva = restore_seva(db, seva_id)
    if not updated_seva:
        raise HTTPException(status_code=404, detail="Seva not found")
    return updated_seva


@app.delete("/sevas/{seva_id}/permanent", status_code=status.HTTP_204_NO_CONTENT, tags=["Seva Catalog"])
def permanently_delete_seva_endpoint(
    seva_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Permanently delete a soft-deleted Seva (Admin only).
    The seva must already be inactive (in recycle bin).
    """
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can permanently delete Sevas"
        )
    
    from app.crud import permanently_delete_seva
    try:
        result = permanently_delete_seva(db, seva_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not result:
        raise HTTPException(status_code=404, detail="Seva not found")
    return None


@app.delete("/sevas/recycle-bin/empty", status_code=status.HTTP_200_OK, tags=["Seva Catalog"])
def empty_recycle_bin_endpoint(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Permanently delete ALL inactive sevas (Admin only). Empties the recycle bin.
    """
    if current_user.role.lower() != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Admins can empty the recycle bin"
        )
    
    from app.crud import permanently_delete_all_inactive_sevas
    count = permanently_delete_all_inactive_sevas(db)
    return {"message": f"Permanently deleted {count} seva(s)", "count": count}



# =============================================================================
# API Routes - Booking / Transactions
# =============================================================================

@app.get("/reports", tags=["Reports"])
def get_reports(start_date: str, end_date: str, db: Session = Depends(get_db)):
    """
    Enhanced financial report with comparison period, hourly heatmap,
    ATV, cash/UPI splits, and shaswata count.
    """
    try:
        report = get_enhanced_report(db, start_date, end_date)
        report["period"] = {"start": start_date, "end": end_date}
        return report
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/collection", tags=["Reports"])
def get_report_collection(start_date: str, end_date: str, db: Session = Depends(get_db)):
    """
    Get line-item transaction details for exports and drill-down.
    """
    try:
        txns = get_collection_details(db, start_date, end_date)
        return {"transactions": txns, "count": len(txns)}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/export/excel", tags=["Reports"])
def export_report_excel(start_date: str, end_date: str, db: Session = Depends(get_db)):
    """
    Export a multi-tab Excel report (.xlsx) for the given date range.
    Tabs: Summary, Transactions, Seva Breakdown.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as excel_io

        report = get_enhanced_report(db, start_date, end_date)
        txns = get_collection_details(db, start_date, end_date)

        wb = Workbook()

        # --- Tab 1: Summary ---
        ws1 = wb.active
        ws1.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        bold = Font(bold=True)

        ws1.append(["SHREE SUBRAMANYA TEMPLE — Financial Report"])
        ws1.append([f"Period: {start_date} to {end_date}"])
        ws1.append([])
        ws1.append(["Metric", "Current Period", "Previous Period", "Change %"])
        for cell in ws1[4]:
            cell.font = header_font
            cell.fill = header_fill

        fin = report["financials"]
        comp = report["comparison"]
        ws1.append(["Total Collection", fin["total"], comp["prev_total"], f"{comp['total_change']}%"])
        ws1.append(["Booking Count", fin["tx_count"], comp["prev_count"], f"{comp['count_change']}%"])
        ws1.append(["Avg. Ticket Value", fin["atv"], comp["prev_atv"], f"{comp['atv_change']}%"])
        ws1.append(["Cash Collection", fin["cash"], "", f"{fin['cash_pct']}%"])
        ws1.append(["UPI Collection", fin["upi"], "", f"{fin['upi_pct']}%"])
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 20
        ws1.column_dimensions['D'].width = 15

        # --- Tab 2: Transactions ---
        ws2 = wb.create_sheet(title="Transactions")
        ws2.append(["Receipt #", "Devotee", "Seva", "Amount", "Mode", "Time"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        for t in txns:
            ws2.append([t["receipt_no"], t["devotee_name"], t["seva_name"],
                        t["amount"], t["payment_mode"], t["time"]])
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 25
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 20

        # --- Tab 3: Seva Breakdown ---
        ws3 = wb.create_sheet(title="Seva Breakdown")
        ws3.append(["Seva Name", "Count", "Revenue"])
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
        for s in report["seva_stats"]:
            ws3.append([s["name"], s["count"], s["revenue"]])
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 15

        # Save to bytes
        output = excel_io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Temple_Report_{start_date}_to_{end_date}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/book-seva", response_model=TransactionResponse, tags=["Booking"])
def book_seva(transaction: TransactionCreate, db: Session = Depends(get_db)):
    # Security Scan (SQL Sentinel & XSS)
    validate_transaction_payload(transaction)
    
    try:
        return book_seva_transaction(db=db, transaction=transaction)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Booking failed: {str(e)}")

@app.get("/transactions", tags=["Transactions"])
def list_transactions(
    date: Optional[str] = None,
    payment_mode: Optional[str] = None,
    seva_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 200,
    sort_by: str = "time_desc",
    db: Session = Depends(get_db)
):
    """
    Get paginated, filterable transactions for a specific date.
    
    - **date**: YYYY-MM-DD (defaults to today)
    - **payment_mode**: CASH or UPI
    - **seva_id**: Filter by specific seva
    - **skip/limit**: Pagination
    - **sort_by**: time_desc, time_asc, amount_desc, amount_asc, name_asc
    """
    return get_daily_transactions(
        db, date=date, payment_mode=payment_mode, seva_id=seva_id,
        skip=skip, limit=limit, sort_by=sort_by
    )


@app.get("/transactions/stats", tags=["Transactions"])
def transaction_stats(date: Optional[str] = None, db: Session = Depends(get_db)):
    """
    Get aggregate statistics for a date: totals, payment breakdown,
    seva-wise breakdown, and hourly trend data for charts.
    """
    return get_daily_stats(db, date)


@app.get("/transactions/export", tags=["Transactions"])
def export_transactions(date: Optional[str] = None, db: Session = Depends(get_db)):
    """Export transactions for a date as CSV download."""
    from fastapi.responses import StreamingResponse
    import csv, io
    
    data = get_daily_transactions(db, date=date, limit=10000)
    transactions = data["transactions"]
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Receipt #", "Devotee Name", "Seva", "Amount (₹)", "Payment Mode", "Time"])
    
    for t in transactions:
        time_str = str(t.get("transaction_date", ""))
        writer.writerow([
            t.get("receipt_no", ""),
            t.get("devotee_name", ""),
            t.get("seva_name", ""),
            t.get("amount_paid", 0),
            t.get("payment_mode", ""),
            time_str
        ])
    
    output.seek(0)
    filename = f"transactions_{date or 'today'}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.put("/transactions/{transaction_id}/note", tags=["Transactions"])
def update_transaction_note(transaction_id: int, note: str = "", db: Session = Depends(get_db)):
    """Update the note on a transaction (inline edit from list)."""
    from sqlalchemy import text as sa_text
    result = db.execute(
        sa_text("SELECT id FROM transactions WHERE id = :id"),
        {"id": transaction_id}
    ).fetchone()
    if not result:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    db.execute(
        sa_text("UPDATE transactions SET notes = :note WHERE id = :id"),
        {"id": transaction_id, "note": note}
    )
    db.commit()
    return {"message": "Note updated", "id": transaction_id, "note": note}


# =============================================================================
# API Routes - Devotee Management (Auto-Fill) -- NEW ADDITION
# =============================================================================

@app.get("/devotees/{phone}", tags=["Devotees"])
def get_devotee_details(phone: str, db: Session = Depends(get_db)):
    """Fetch devotee details by phone number for Auto-Fill (bilingual)."""
    query = text("""
        SELECT full_name_en, full_name_kn, gothra_en, gothra_kn, nakshatra, rashi, area, pincode
        FROM devotees WHERE phone_number = :phone
    """)
    result = db.execute(query, {"phone": phone}).fetchone()
    
    if not result:
        raise HTTPException(status_code=404, detail="Devotee not found")
        
    return {
        "full_name": result[0],       # Backward compat
        "full_name_en": result[0],
        "full_name_kn": result[1],
        "gothra": result[2],          # Backward compat
        "gothra_en": result[2],
        "gothra_kn": result[3],
        "nakshatra": result[4],
        "rashi": result[5],
        "area": result[6] if len(result) > 6 else None,
        "pincode": result[7] if len(result) > 7 else None
    }

# =============================================================================
# Security Helpers (SQL Sentinel & XSS Filter)
# =============================================================================
import re

def validate_safe_input(text: str, field_name: str):
    """
    rejects payloads containing potential XSS or SQL Injection vectors.
    """
    if not text:
        return
    
    # 1. XSS / Malicious Script Check
    # Triggers on <script>, javascript:, variables with on<event> (onload, onerror), iframes
    xss_pattern = r"<script|javascript:|onload=|onerror=|<iframe>|<object>|data:text/html"
    if re.search(xss_pattern, text, re.IGNORECASE):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail=f"Security Alert: Malicious content detected in '{field_name}'"
        )
        
    # 2. SQL Injection Sentinel (Basic Pattern Match)
    # Strict checks for specific fields like Gothra/Nakshatra where code shouldn't be present
    # Blocking semi-colons and SQL comment dashes
    if field_name.lower() in ["gothra", "nakshatra", "rashi"]:
        sql_pattern = r"[;']|--"
        if re.search(sql_pattern, text):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, 
                detail=f"Security Alert: Invalid characters detected in '{field_name}'"
            )

def validate_transaction_payload(data):
    """Scan all string fields in the payload"""
    if hasattr(data, "model_dump"):
        data = data.model_dump()
    elif hasattr(data, "dict"):
        data = data.dict()
    
    for key, value in data.items():
        if isinstance(value, str):
            validate_safe_input(value, key)

# =============================================================================
# API Routes - Shaswata (Perpetual Puja)
# =============================================================================

@app.post("/shaswata/subscribe", response_model=ShaswataResponse, tags=["Shaswata"])
def subscribe_shaswata(subscription: ShaswataCreate, db: Session = Depends(get_db)):
    # Security Scan
    validate_transaction_payload(subscription)
    
    try:
        return create_shaswata_subscription(db=db, subscription=subscription)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Subscription failed: {str(e)}")

@app.get("/shaswata/subscriptions", tags=["Shaswata"])
def list_shaswata_subscriptions(active_only: bool = True, db: Session = Depends(get_db)):
    return get_shaswata_subscriptions(db, active_only=active_only)

@app.get("/shaswata/list", tags=["Shaswata"])
def list_shaswata_alias(active_only: bool = True, db: Session = Depends(get_db)):
    """Alias for /shaswata/subscriptions for frontend compatibility."""
    return get_shaswata_subscriptions(db, active_only=active_only)


@app.get("/shaswata/tomorrow", tags=["Shaswata"])
def get_tomorrow_shaswata_pujas(db: Session = Depends(get_db)):
    """
    Get all Shaswata subscriptions scheduled for TOMORROW.
    Used to power the notification bell in the Navbar.
    Returns list of pujas with devotee details and address confirmation status.
    """
    from app.models import Devotee
    tomorrow = date.today() + timedelta(days=1)
    
    pc = PanchangCalculator()
    panchangam_tomorrow = pc.calculate(tomorrow)
    
    # 1. LUNAR subscriptions matching tomorrow's Tithi
    lunar_query = text("""
        SELECT ss.id, d.id as devotee_id, d.full_name_en, d.phone_number, d.address, d.area, d.pincode,
               sc.name_eng, ss.maasa, ss.paksha, ss.tithi, ss.occasion,
               d.address_confirmed, d.address_confirmation_sent_at
        FROM shaswata_subscriptions ss
        JOIN devotees d ON ss.devotee_id = d.id
        JOIN seva_catalog sc ON ss.seva_id = sc.id
        WHERE ss.is_active = TRUE AND ss.subscription_type = 'LUNAR'
        AND ss.maasa = :maasa AND ss.paksha = :paksha AND ss.tithi = :tithi
    """)
    lunar_result = db.execute(lunar_query, {
        "maasa": panchangam_tomorrow["attributes"]["maasa"],
        "paksha": panchangam_tomorrow["attributes"]["paksha"],
        "tithi": panchangam_tomorrow["attributes"]["tithi"]
    }).fetchall()
    
    # 2. GREGORIAN subscriptions matching tomorrow's date
    greg_query = text("""
        SELECT ss.id, d.id as devotee_id, d.full_name_en, d.phone_number, d.address, d.area, d.pincode,
               sc.name_eng, ss.event_day, ss.event_month, ss.occasion,
               d.address_confirmed, d.address_confirmation_sent_at
        FROM shaswata_subscriptions ss
        JOIN devotees d ON ss.devotee_id = d.id
        JOIN seva_catalog sc ON ss.seva_id = sc.id
        WHERE ss.is_active = TRUE AND ss.subscription_type = 'GREGORIAN'
        AND ss.event_day = :day AND ss.event_month = :month
    """)
    greg_result = db.execute(greg_query, {"day": tomorrow.day, "month": tomorrow.month}).fetchall()
    
    pujas = []
    for row in lunar_result:
        pujas.append({
            "id": row[0], "devotee_id": row[1], "name": row[2], "phone": row[3],
            "address": row[4], "area": row[5], "pincode": row[6],
            "seva": row[7], "date_info": f"{row[8]} {row[9]} {row[10]}",
            "occasion": row[11], "type": "LUNAR",
            "address_confirmed": row[12] or False,
            "confirmation_sent": row[13] is not None
        })
    for row in greg_result:
        month_names = ["", "January", "February", "March", "April", "May", "June",
                       "July", "August", "September", "October", "November", "December"]
        month_idx = row[9]
        month_str = month_names[month_idx] if 0 < month_idx <= 12 else str(month_idx)
        pujas.append({
            "id": row[0], "devotee_id": row[1], "name": row[2], "phone": row[3],
            "address": row[4], "area": row[5], "pincode": row[6],
            "seva": row[7], "date_info": f"{month_str} {row[8]}",
            "occasion": row[10], "type": "GREGORIAN",
            "address_confirmed": row[11] or False,
            "confirmation_sent": row[12] is not None
        })
    
    return {
        "count": len(pujas),
        "date": tomorrow.strftime("%d-%m-%Y"),
        "pujas": pujas
    }


@app.patch("/devotees/{devotee_id}/send-address-confirmation", tags=["Shaswata"])
def send_address_confirmation_endpoint(devotee_id: int, db: Session = Depends(get_db)):
    """
    Mark that an address confirmation WhatsApp message was sent to a devotee.
    """
    try:
        devotee = send_address_confirmation(db, devotee_id)
        return {
            "status": "confirmation_sent",
            "devotee_id": devotee.id,
            "name": devotee.full_name_en,
            "sent_at": str(devotee.address_confirmation_sent_at)
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.patch("/devotees/{devotee_id}/confirm-address", tags=["Shaswata"])
def confirm_address_endpoint(
    devotee_id: int,
    address: Optional[str] = None,
    area: Optional[str] = None,
    pincode: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Confirm a devotee's address. Optionally update if changed.
    """
    try:
        devotee = confirm_devotee_address(db, devotee_id, address, area, pincode)
        return {
            "status": "confirmed",
            "devotee_id": devotee.id,
            "name": devotee.full_name_en,
            "address": devotee.address,
            "area": devotee.area,
            "pincode": devotee.pincode,
            "confirmed_at": str(devotee.address_confirmed_at)
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


# =============================================================================
# API Routes - Dispatch & Feedback Automation (Stage 2)
# =============================================================================

@app.patch("/subscriptions/{subscription_id}/dispatch", tags=["Shaswata Automation"])
def mark_subscription_dispatched(subscription_id: int, db: Session = Depends(get_db)):
    """
    Log that prasadam has been dispatched for this subscription.
    Sets last_dispatch_date to today.
    """
    try:
        return log_dispatch(db, subscription_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Dispatch logging failed: {str(e)}")


@app.patch("/subscriptions/{subscription_id}/feedback", tags=["Shaswata Automation"])
def mark_feedback_sent(subscription_id: int, db: Session = Depends(get_db)):
    """
    Log that feedback message has been sent for this subscription.
    Sets last_feedback_date to today.
    """
    try:
        return log_feedback_sent(db, subscription_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Feedback logging failed: {str(e)}")


@app.get("/subscriptions/pending-feedback", tags=["Shaswata Automation"])
def get_subscriptions_pending_feedback(db: Session = Depends(get_db)):
    """
    Get all subscriptions that need feedback (5+ days after dispatch).
    Returns list of devotees who should be contacted for prasadam confirmation.
    """
    try:
        return get_pending_feedback_subscriptions(db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Query failed: {str(e)}")

# =============================================================================
# API Routes - Shaswata Phase 2 (Event Lifecycle & Messaging)
# =============================================================================

@app.post("/shaswata/events/populate", tags=["Shaswata Manager"])
def populate_shaswata_events(days: int = 30, db: Session = Depends(get_db)):
    """
    Populate shaswata_events for the next N days.
    Scans all active subscriptions and creates PENDING events.
    Idempotent: safe to call multiple times.
    """
    try:
        result = populate_upcoming_events(db, days_ahead=days)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Event population failed: {str(e)}")


@app.get("/shaswata/upcoming", tags=["Shaswata Manager"])
def get_upcoming_shaswata(days: int = 30, db: Session = Depends(get_db)):
    """
    Get all upcoming Shaswata events for the admin dashboard.
    Returns enriched data with devotee details and lifecycle status.
    """
    try:
        events = get_upcoming_events(db, days_ahead=days)
        return {
            "count": len(events),
            "events": events
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Query failed: {str(e)}")


@app.post("/shaswata/events/{event_id}/dispatch", tags=["Shaswata Manager"])
def dispatch_event(
    event_id: int,
    dispatch_ref: Optional[str] = None,
    dispatch_method: str = "POST",
    db: Session = Depends(get_db)
):
    """
    Mark a specific event as DISPATCHED with optional tracking reference.
    Also sends a notification to the devotee.
    """
    try:
        return mark_event_dispatched(db, event_id, dispatch_ref, dispatch_method)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Dispatch failed: {str(e)}")


@app.post("/shaswata/events/{event_id}/delivery-feedback", tags=["Shaswata Manager"])
def submit_delivery_feedback(
    event_id: int,
    received: bool = True,
    notes: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Record delivery feedback from a devotee.
    received=True  → Status = DELIVERED (happy path)
    received=False → Status = NOT_RECEIVED or ADDRESS_CHANGED (follow-up needed)
    """
    try:
        return record_delivery_feedback(db, event_id, received, notes)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Feedback recording failed: {str(e)}")


@app.get("/shaswata/pending-delivery-checks", tags=["Shaswata Manager"])
def pending_delivery_checks(db: Session = Depends(get_db)):
    """
    Get events dispatched 4+ days ago that haven't been checked for delivery.
    Admin should send 'Did you receive Prasadam?' message to these devotees.
    """
    try:
        checks = get_pending_delivery_checks(db)
        return {"count": len(checks), "pending": checks}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Query failed: {str(e)}")


@app.post("/shaswata/dispatch-adhoc", tags=["Shaswata Manager"])
def legacy_dispatch_adapter(
    payload: ShaswataDispatchAdhoc,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Unified Dispatch Endpoint.
    Handles ad-hoc dispatch for manual 'Send Prasadam' actions in Search/Calendar view.
    Ensures event exists -> Marks Dispatched -> Logs Communication.
    """
    if current_user.role.lower() not in ["admin", "clerk"]:
        raise HTTPException(status_code=403, detail="Not authorized")
        
    try:
        return dispatch_adhoc_event(
            db, 
            payload.subscription_id, 
            payload.dispatch_date, 
            payload.dispatch_ref, 
            payload.dispatch_method
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/shaswata/events/{event_id}/send-reminder", tags=["Shaswata Manager"])
def send_event_reminder(event_id: int, db: Session = Depends(get_db)):
    """
    Send a 'Pooja Tomorrow' reminder for a specific event.
    V1: Logs to console + communication_logs table.
    """
    try:
        return MessagingService.send_reminder(db, event_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Reminder failed: {str(e)}")


@app.post("/shaswata/events/{event_id}/send-delivery-check", tags=["Shaswata Manager"])
def send_delivery_check_message(event_id: int, db: Session = Depends(get_db)):
    """
    Send a 'Did you receive Prasadam?' message for a dispatched event.
    """
    try:
        return MessagingService.send_delivery_check(db, event_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delivery check failed: {str(e)}")


@app.get("/shaswata/communication-history", tags=["Shaswata Manager"])
def get_comm_history(
    devotee_id: Optional[int] = None,
    subscription_id: Optional[int] = None,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """
    Get communication history for a devotee or subscription.
    Used in the admin detail panel to see all messages sent.
    """
    try:
        history = get_communication_history(db, devotee_id, subscription_id, limit)
        return {"count": len(history), "messages": history}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Query failed: {str(e)}")

# =============================================================================
# API Routes - Priest Dashboard
# =============================================================================

@app.get("/daily-sankalpa", tags=["Priest Dashboard"])
def get_daily_sankalpa(date_str: str = None, lang: str = "en", db: Session = Depends(get_db)):
    """
    Get daily sankalpa/schedule with cached Panchangam. 
    If date_str provided (DD-MM-YYYY), use that date. 
    Otherwise default to today.
    """
    import json as json_lib
    import hashlib
    
    try:
        if date_str:
            target_date = datetime.strptime(date_str, "%d-%m-%Y").date()
        else:
            target_date = date.today()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use DD-MM-YYYY")

    # ── Read Location & Ayanamsa from SystemSettings ──
    def _get_setting(key, default):
        s = db.query(SystemSetting).filter_by(key=key).first()
        return s.value if s else default
    
    temple_lat = _get_setting("temple_lat", "12.6745")
    temple_lon = _get_setting("temple_lon", "75.3370")
    temple_elev = _get_setting("temple_elevation", "120")
    ayanamsa = _get_setting("panchang_ayanamsa", "lahiri")
    cache_version = int(_get_setting("panchang_cache_version", "1"))
    
    location_hash = hashlib.md5(f"{temple_lat}:{temple_lon}:{ayanamsa}".encode()).hexdigest()[:12]

    # ── Cache Check ──
    cached = db.query(DailyPanchang).filter_by(date=target_date).first()
    if cached and cached.version == cache_version and cached.location_hash == location_hash:
        panchangam = json_lib.loads(cached.data_json)
    else:
        # Calculate fresh
        pc = PanchangCalculator(
            lat=temple_lat,
            lon=temple_lon,
            elevation=temple_elev,
            ayanamsa=ayanamsa
        )
        panchangam = pc.calculate(target_date, lang=lang)
        
        # Save to cache
        try:
            if cached:
                cached.data_json = json_lib.dumps(panchangam, ensure_ascii=False)
                cached.version = cache_version
                cached.location_hash = location_hash
            else:
                new_cache = DailyPanchang(
                    date=target_date,
                    data_json=json_lib.dumps(panchangam, ensure_ascii=False),
                    version=cache_version,
                    location_hash=location_hash
                )
                db.add(new_cache)
            db.commit()
        except Exception as cache_err:
            print(f"[WARN] Panchang cache write failed: {cache_err}")
            db.rollback()
    
    # 1. Lunar Query - Find subscriptions matching TODAY's Tithi
    lunar_query = text("""
        SELECT ss.id, d.full_name_en, d.phone_number, d.gothra_en, sc.name_eng, 
               ss.maasa, ss.paksha, ss.tithi, ss.notes, d.address, d.nakshatra, d.rashi, ss.occasion
        FROM shaswata_subscriptions ss
        JOIN devotees d ON ss.devotee_id = d.id
        JOIN seva_catalog sc ON ss.seva_id = sc.id
        WHERE ss.is_active = TRUE 
        AND ss.subscription_type = 'LUNAR'
        AND ss.maasa = :maasa
        AND ss.paksha = :paksha
        AND ss.tithi = :tithi
    """)
    
    lunar_result = db.execute(lunar_query, {
        "maasa": panchangam["attributes"]["maasa"],
        "paksha": panchangam["attributes"]["paksha"],
        "tithi": panchangam["attributes"]["tithi"]
    }).fetchall()
    
    lunar_pujas = [{
        "id": row[0], "name": row[1], "phone": row[2], "gothra": row[3],
        "seva": row[4], "date_info": f"{row[5]} {row[6]} {row[7]}", "notes": row[8], "address": row[9],
        "nakshatra": row[10], "rashi": row[11], "type": "LUNAR", "occasion": row[12]
    } for row in lunar_result]
    
    # 2. Gregorian Query (Dynamic based on selected date)
    gregorian_query = text("""
        SELECT ss.id, d.full_name_en, d.phone_number, d.gothra_en, sc.name_eng, 
               ss.event_day, ss.event_month, ss.notes, d.address, d.nakshatra, d.rashi, ss.occasion
        FROM shaswata_subscriptions ss
        JOIN devotees d ON ss.devotee_id = d.id
        JOIN seva_catalog sc ON ss.seva_id = sc.id
        WHERE ss.is_active = TRUE AND ss.subscription_type = 'GREGORIAN'
          AND ss.event_day = :day AND ss.event_month = :month
    """)
    gregorian_result = db.execute(gregorian_query, {"day": target_date.day, "month": target_date.month}).fetchall()
    
    gregorian_pujas = []
    month_names = ["", "January", "February", "March", "April", "May", "June", 
                  "July", "August", "September", "October", "November", "December"]

    for row in gregorian_result:
        month_idx = row[6]
        month_str = month_names[month_idx] if 0 < month_idx <= 12 else str(month_idx)
        
        gregorian_pujas.append({
            "id": row[0], "name": row[1], "phone": row[2], "gothra": row[3],
            "seva": row[4], "date_info": f"{month_str} {row[5]}", "notes": row[7], "address": row[8],
            "nakshatra": row[9], "rashi": row[10], "type": "GREGORIAN", "occasion": row[11]
        })

    # 3. One-Time Transactions (Check BOTH seva_date AND transaction_date)
    transaction_query = text("""
        SELECT t.id, t.devotee_name, d.phone_number, d.gothra_en, s.name_eng, t.notes, d.address, d.nakshatra, d.rashi
        FROM transactions t
        JOIN devotees d ON t.devotee_id = d.id
        JOIN seva_catalog s ON t.seva_id = s.id
        WHERE t.seva_date = :date 
           OR CAST(t.transaction_date AS DATE) = :date
    """)
    transaction_result = db.execute(transaction_query, {"date": target_date}).fetchall()

    transaction_pujas = [{
        "id": row[0], "name": row[1], "phone": row[2], "gothra": row[3],
        "seva": row[4], "date_info": "One-Time", "notes": row[5] or "Booked via App", "address": row[6],
        "nakshatra": row[7], "rashi": row[8], "type": "BOOKING"
    } for row in transaction_result]
    
    # 4. Calculate Daily Revenue
    revenue_query = text("""
        SELECT SUM(amount_paid) FROM transactions 
        WHERE seva_date = :date OR DATE(transaction_date) = :date
    """)
    daily_revenue = db.execute(revenue_query, {"date": target_date}).scalar() or 0

    return {
        "date": {"day": target_date.day, "month": target_date.month, "year": target_date.year, "weekday": target_date.strftime("%A")},
        "panchangam": panchangam,
        "pujas": lunar_pujas + gregorian_pujas + transaction_pujas,
        "revenue": daily_revenue,
        "festivals": [panchangam["is_festival"]] if panchangam.get("is_festival") else (["Shiva Rathri (Upcoming)", "Pradosha"] if panchangam["attributes"]["maasa"] == "Magha" else ["Daily Sevas"])
    }

# =============================================================================
# API Routes - Financial Reports (CLEAN VERSION)
# =============================================================================

@app.get("/reports/collection", tags=["Financial Reports"])
def get_collection_report(start_date: str = None, end_date: str = None, db: Session = Depends(get_db)):
    try:
        # Parse DD-MM-YYYY format
        start = datetime.strptime(start_date, "%d-%m-%Y").date() if start_date else date.today()
        end = datetime.strptime(end_date, "%d-%m-%Y").date() if end_date else start
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use DD-MM-YYYY")
    
    try:
        query = text("""
            SELECT t.id, t.receipt_no, d.full_name_en, d.phone_number, sc.name_eng, 
                   t.amount_paid, t.payment_mode, t.transaction_date, t.notes
            FROM transactions t
            JOIN devotees d ON t.devotee_id = d.id
            JOIN seva_catalog sc ON t.seva_id = sc.id
            WHERE CAST(t.transaction_date AS DATE) >= :start AND CAST(t.transaction_date AS DATE) <= :end
            ORDER BY t.transaction_date DESC
        """)
        result = db.execute(query, {"start": start, "end": end}).fetchall()
        
        transactions = []
        total_cash, total_upi = 0, 0
        
        for row in result:
            amt = float(row[5] or 0)
            mode = str(row[6] or "").upper()
            if mode == 'CASH': total_cash += amt
            elif mode == 'UPI': total_upi += amt
            
            transactions.append({
                "id": row[0], "receipt_no": row[1], "devotee_name": row[2], "phone": row[3],
                "seva_name": row[4], "amount": amt, "payment_mode": mode,
                "time": row[7].strftime("%Y-%m-%d %I:%M %p") if row[7] else "", "notes": row[8]
            })
            
        return {
            "start_date": start, "end_date": end,
            "summary": {"cash": total_cash, "upi": total_upi, "total": total_cash + total_upi},
            "transactions": transactions
        }
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(status_code=500, detail="Database error")

@app.get("/reports/export", tags=["Financial Reports"])
def export_report(start_date: str = None, end_date: str = None, db: Session = Depends(get_db)):
    try:
        # Standardize on YYYY-MM-DD from frontend
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else start
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export Failed: {str(e)}")

    # Create Excel Workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    
    # --- SHEET 1: SUMMARY ---
    ws_summary = wb.active
    ws_summary.title = "Financial Summary"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid") # Saffron
    subheader_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid") # Slate
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Title
    ws_summary.merge_cells('A1:D2')
    cell = ws_summary['A1']
    cell.value = "SHREE SUBRAMANYA TEMPLE - FINANCIAL REPORT"
    cell.font = Font(bold=True, size=16, color="F97316")
    cell.alignment = center_align

    ws_summary['A3'] = f"Period: {start} to {end}"
    ws_summary['A3'].font = Font(italic=True)

    # 1. Totals Table
    ws_summary['A5'] = "FINANCIAL OVERVIEW"
    ws_summary['A5'].font = Font(bold=True)
    
    headers = ["Category", "Amount"]
    ws_summary.append([]) # Gap
    ws_summary.append(headers)
    
    # Style Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=6, column=col_num)
        cell.fill = subheader_fill
        cell.font = header_font
        cell.alignment = center_align

    # Calculate Data
    stmt = text("""
        SELECT 
            COALESCE(SUM(amount_paid), 0),
            COALESCE(SUM(CASE WHEN payment_mode = 'CASH' THEN amount_paid ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN payment_mode = 'UPI' THEN amount_paid ELSE 0 END), 0)
        FROM transactions
        WHERE DATE(transaction_date) >= :start AND DATE(transaction_date) <= :end
    """)
    totals = db.execute(stmt, {"start": start, "end": end}).fetchone()
    total_val, cash_val, upi_val = totals[0], totals[1], totals[2]

    ws_summary.append(["Total Collection", total_val])
    ws_summary.append(["Cash", cash_val])
    ws_summary.append(["UPI", upi_val])

    # 2. Seva Breakdown
    ws_summary['D5'] = "TOP SEVAS"
    ws_summary['D5'].font = Font(bold=True)
    
    headers_seva = ["Seva Name", "Count", "Revenue"]
    # Manually place headers at D6
    for i, h in enumerate(headers_seva):
        cell = ws_summary.cell(row=6, column=4+i)
        cell.value = h
        cell.fill = subheader_fill
        cell.font = header_font
        cell.alignment = center_align

    seva_stmt = text("""
        SELECT s.name_eng, COUNT(t.id), SUM(t.amount_paid)
        FROM transactions t
        JOIN seva_catalog s ON t.seva_id = s.id
        WHERE CAST(t.transaction_date AS DATE) >= :start AND CAST(t.transaction_date AS DATE) <= :end
        GROUP BY s.name_eng
        ORDER BY SUM(t.amount_paid) DESC
        LIMIT 10
    """)
    sevas = db.execute(seva_stmt, {"start": start, "end": end}).fetchall()
    
    for idx, row in enumerate(sevas, 7):
        ws_summary.cell(row=idx, column=4).value = row[0]
        ws_summary.cell(row=idx, column=5).value = row[1]
        ws_summary.cell(row=idx, column=6).value = row[2]

    # Adjust Column Widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['D'].width = 30
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 20

    # --- SHEET 2: DETAILED TRANSACTIONS ---
    ws_detail = wb.create_sheet(title="Detailed Transactions")
    
    headers_detail = ["Receipt No", "Date", "Devotee Name", "Seva Name", "Mode", "Amount", "Notes"]
    ws_detail.append(headers_detail)
    
    # Style Header
    for col_num, _ in enumerate(headers_detail, 1):
        cell = ws_detail.cell(row=1, column=col_num)
        cell.fill = header_fill # Saffron
        cell.font = header_font
        cell.alignment = center_align

    # Fetch Data
    tx_stmt = text("""
        SELECT t.receipt_no, strftime('%d-%m-%Y %I:%M %p', t.transaction_date), 
               d.full_name_en, sc.name_eng, t.payment_mode, t.amount_paid, t.notes
        FROM transactions t
        JOIN devotees d ON t.devotee_id = d.id
        JOIN seva_catalog sc ON t.seva_id = sc.id
        WHERE CAST(t.transaction_date AS DATE) >= :start AND CAST(t.transaction_date AS DATE) <= :end
        ORDER BY t.transaction_date DESC
    """)
    transactions = db.execute(tx_stmt, {"start": start, "end": end}).fetchall()
    
    current_date_str = None
    date_header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue
    
    for row in transactions:
        # Extract Date info (DD-MM-YYYY)
        # row[1] is "DD-MM-YYYY HH:MM PM"
        # We want just the date part for grouping
        date_part = row[1].split(' ')[0] if row[1] else "Unknown Date"
        
        # Check for Group Change
        if date_part != current_date_str:
            current_date_str = date_part
            # Insert Group Header
            ws_detail.append([f"Date: {current_date_str}"])
            # Merge Cells for Header
            current_row = ws_detail.max_row
            ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            # Style Header
            header_cell = ws_detail.cell(row=current_row, column=1)
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.fill = date_header_fill
            header_cell.alignment = Alignment(horizontal='left')

        # Insert Data Row
        ws_detail.append([
            row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        ])

    # Auto-width
    for col in ws_detail.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_detail.column_dimensions[column].width = adjusted_width

    # Save to IO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Temple_Report_{start}_{end}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# =============================================================================
# Level 13: The Legacy Eraser Endpoint
# =============================================================================
from fastapi import UploadFile, File
import shutil
import os
from app.legacy_migrator import migrate_legacy_data

@app.post("/admin/migrate", tags=["Admin Operations"])
def migrate_legacy_database(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Drag & Drop Portal for Legacy Data (CSV/PDF).
    Migrates 2010-era records into the modern Transaction table.
    Supports:
    - CSV: Standard headers
    - PDF: Tables extracted automatically
    """
    if current_user.role.lower() != "admin":
        raise HTTPException(status_code=403, detail="Only Admins can rewrite history.")

    # Save temp file
    temp_filename = f"temp_{file.filename}"
    with open(temp_filename, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        # Invoke the heuristic engine
        # Note: migrate_legacy_data currently handles its own DB session, 
        # but we should ideally pass the 'db' session. 
        # For now, running it as a separate process logic or function call is fine.
        migrate_legacy_data(temp_filename)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(temp_filename):
            os.remove(temp_filename)

# =============================================================================
# Level 16: The Celestial Compass (Reverse Panchangam Search)
# =============================================================================

@app.get("/panchangam/find", tags=["Panchangam"])
def find_date_by_panchangam(
    masa: str, 
    paksha: str, 
    tithi: str, 
    year: int = None
):
    """
    Finds the Gregorian date for a given Panchangam date (Masa, Paksha, Tithi).
    Iterates through the specified Gregorian year to find the first match.
    """
    target_year = year if year else datetime.now().year
    
    # Normalize inputs
    t_masa = masa.lower().strip()
    t_paksha = paksha.lower().strip()
    t_tithi = tithi.lower().strip()
    
    # Alias handling
    # Calculator uses 'Shasthi' (one 'h' after 's'), but inputs might be 'Shashthi' or 'Shasti'
    if 'shashthi' in t_tithi: t_tithi = 'shasthi'
    if 'shasti' in t_tithi: t_tithi = 'shasthi'

    # Optimization: Reduce search space based on Masa
    # (Optional, but simple iteration is fast enough for 365 days)
    
    start_date = date(target_year, 1, 1)
    # Search for 380 days to cover overlaps into next year
    for i in range(380):
        current_date = start_date + timedelta(days=i)
        
        # Calculate Panchangam for this date
        p = PanchangCalculator.calculate(current_date)
        attrs = p['attributes']
        
        # Check Match
        # Case-insensitive partial matches
        p_masa = attrs['maasa'].lower()
        p_paksha = attrs['paksha'].lower()
        p_tithi = attrs['tithi'].lower()
        
        # We check if target input is IN the calculated string (e.g. "Shashthi" in "Shashthi")
        if t_masa in p_masa and t_paksha in p_paksha and t_tithi in p_tithi:
            return {"date": p['date'], "panchangam": attrs}
            
            
    raise HTTPException(status_code=404, detail="Panchangam date not found in this year")

# =============================================================================
# Level 17: The Divine Scroll (Thermal Printer Integration)
# =============================================================================
from app.printer_service import generate_receipt_image, print_receipt_image

from fastapi.responses import FileResponse
import os

@app.post("/print/preview", tags=["Device Integration"])
def preview_receipt(data: dict):
    """
    Generates a receipt image and returns it for preview without printing.
    """
    try:
        # Generate Image (Devotee Copy as default preview)
        data['copy_label'] = "** PREVIEW **"
        image_path = generate_receipt_image(data)
        
        if os.path.exists(image_path):
            return FileResponse(image_path, media_type="image/jpeg")
        else:
            raise HTTPException(status_code=500, detail="Failed to generate preview image")
    except Exception as e:
        print(f"Preview Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/print/receipt", tags=["Device Integration"])
def print_receipt(data: dict):
    """
    Generates and prints a thermal receipt for the given data.
    Usage: Payload must match the keys expected by printer_service.
    """
    image_paths = []
    try:
        # 1. Devotee Copy
        data_devotee = data.copy()
        data_devotee['copy_label'] = "** DEVOTEE COPY **"
        image_path_d = generate_receipt_image(data_devotee)
        image_paths.append(image_path_d)
        s1 = print_receipt_image(image_path_d)
        
        # 2. Priest Copy (Archaka)
        data_priest = data.copy()
        data_priest['copy_label'] = "** ARCHAKA COPY **"
        image_path_p = generate_receipt_image(data_priest)
        image_paths.append(image_path_p)
        s2 = print_receipt_image(image_path_p)
        
        # 3. Cleanup temp files
        for path in image_paths:
            if os.path.exists(path):
               try:
                   os.remove(path)
               except:
                   pass
        
        # 4. Check print results
        if s1 and s2:
            return {"status": "success", "message": "Receipts sent to printer (2 Copies)"}
        elif s1 or s2:
            return {"status": "partial_success", "message": "One copy failed to print"}
        else:
            raise HTTPException(status_code=500, detail="Failed to print receipts. Check printer connection.")
            
    except HTTPException:
        raise
    except Exception as e:
        print(f"Print Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/print/image", tags=["Device Integration"])
def print_uploaded_image(file: UploadFile = File(...)):
    """
    Directly prints an uploaded image (from Frontend html2canvas).
    Bypasses server-side rendering issues (font/ligatures).
    """
    try:
        # Save temp file
        file_location = f"temp_receipt_{int(datetime.now().timestamp())}.png"
        with open(file_location, "wb+") as file_object:
            shutil.copyfileobj(file.file, file_object)
            
        # Print
        # We print 2 copies for standard flow
        s1 = print_receipt_image(file_location)
        s2 = print_receipt_image(file_location)
        
        # Cleanup
        if os.path.exists(file_location):
            os.remove(file_location)
            
        if s1 and s2:
             return {"status": "success", "message": "Receipts sent to printer (2 Copies)"}
        elif s1 or s2:
             return {"status": "partial_success", "message": "One copy failed to print"}
        else:
            raise HTTPException(status_code=500, detail="Failed to print receipts. Check printer connection.")

    except Exception as e:
        print(f"Print Image Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# Level 99: The Publishing Ritual (Static Files & SPA)
# =============================================================================

# Determine if running as PyInstaller Bundle
if getattr(sys, 'frozen', False):
    # Running as compiled .exe
    base_dir = sys._MEIPASS
    static_dir = os.path.join(base_dir, 'static')
else:
    # Running from source (dev mode)
    # Assume 'star-frontend/dist' is sibling to 'star-backend'
    base_dir = os.path.dirname(os.path.abspath(__file__))
    static_dir = os.path.join(base_dir, '..', 'star-frontend', 'dist')

# Mount Static Files (if they exist)
if os.path.exists(static_dir):
    app.mount("/assets", StaticFiles(directory=os.path.join(static_dir, "assets")), name="assets")
    
    # Catch-All Route for SPA (must be last)
    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        # Check if file exists in static root (e.g. vite.svg, favicon.ico)
        static_file = os.path.join(static_dir, full_path)
        if os.path.isfile(static_file):
            return FileResponse(static_file)
            
        # Otherwise serve index.html for React Router
        return FileResponse(os.path.join(static_dir, "index.html"))
else:
    print(f"INFO: Static files not found at {static_dir}. Run 'npm run build' first for production mode.")

# =============================================================================
# Level 100: The Launch Sequence (Auto-Start for EXE)
# =============================================================================

def open_browser():
    """Open the browser after a short delay to let the server start."""
    if os.environ.get('ELECTRON_MODE') == 'true':
        print("Running in Electron Mode - Browser launch skipped")
        return
        
    import time
    time.sleep(2)  # Wait for server to start
    webbrowser.open("http://127.0.0.1:8000")

if __name__ == "__main__":
    import uvicorn
    
    # Open browser in a separate thread
    threading.Thread(target=open_browser, daemon=True).start()
    
    # Run the server
    print("=" * 60)
    print("  S.T.A.R. - Subramanya Temple App & Registry")
    print("  Server starting at http://127.0.0.1:8000")
    print("  Press Ctrl+C to stop the server")
    print("=" * 60)
    
    uvicorn.run(app, host="127.0.0.1", port=8000, log_level="warning")
